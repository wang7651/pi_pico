"""
生成測試數據檔案
同時建立 CSV 和 Excel 格式
"""

import csv
from datetime import datetime, timedelta
import random

# 嘗試導入 openpyxl（用於 Excel）
try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠️  未安裝 openpyxl，將只生成 CSV 檔案")

def generate_test_data(count=50):
    """
    生成測試數據
    
    Args:
        count: 要生成的數據筆數
        
    Returns:
        list: 包含測試數據的列表
    """
    data = []
    base_time = datetime.now() - timedelta(hours=count//2)  # 從幾小時前開始
    
    # 基礎溫度和濕度（會有變化）
    base_temp = 25.0
    base_humi = 60.0
    
    for i in range(count):
        # 計算時間（每筆數據間隔幾分鐘）
        timestamp = base_time + timedelta(minutes=i * 5)
        timestamp_str = timestamp.strftime('%Y-%m-%d %H:%M:%S')
        
        # 生成溫度（帶有隨機波動）
        temp_variation = random.uniform(-3, 3)
        temperature = round(base_temp + temp_variation + (i % 10) * 0.5, 2)
        
        # 生成濕度（帶有隨機波動）
        humi_variation = random.uniform(-5, 5)
        humidity = round(base_humi + humi_variation + (i % 8) * 0.8, 2)
        
        # 電燈狀態（模擬白天關、晚上開）
        hour = timestamp.hour
        if 6 <= hour <= 18:
            light_status = "關" if random.random() > 0.3 else "開"
        else:
            light_status = "開" if random.random() > 0.3 else "關"
        
        data.append({
            '時間戳記': timestamp_str,
            '電燈狀態': light_status,
            '溫度': temperature,
            '濕度': humidity
        })
    
    return data

def save_to_csv(data, filename='sensor_data.csv'):
    """儲存為 CSV 檔案"""
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        fieldnames = ['時間戳記', '電燈狀態', '溫度', '濕度']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
    print(f"✅ CSV 檔案已建立: {filename}")
    print(f"   包含 {len(data)} 筆數據")

def save_to_excel(data, filename='sensor_data.xlsx'):
    """儲存為 Excel 檔案"""
    if not HAS_OPENPYXL:
        print("❌ 無法建立 Excel 檔案（需要 openpyxl）")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "感測器數據"
    
    # 寫入標題
    headers = ['時間戳記', '電燈狀態', '溫度', '濕度']
    ws.append(headers)
    
    # 設定標題樣式
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # 寫入數據
    for row in data:
        ws.append([
            row['時間戳記'],
            row['電燈狀態'],
            row['溫度'],
            row['濕度']
        ])
    
    # 調整欄寬
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    
    wb.save(filename)
    print(f"✅ Excel 檔案已建立: {filename}")
    print(f"   包含 {len(data)} 筆數據")

def main():
    """主程式"""
    print("=" * 60)
    print(" 測試數據生成工具")
    print("=" * 60)
    print()
    
    # 生成測試數據
    print("📊 生成測試數據...")
    data = generate_test_data(count=50)
    
    # 顯示數據統計
    temps = [d['溫度'] for d in data]
    humis = [d['濕度'] for d in data]
    lights_on = sum(1 for d in data if d['電燈狀態'] == '開')
    
    print(f"\n📈 數據統計:")
    print(f"   總筆數: {len(data)}")
    print(f"   時間範圍: {data[0]['時間戳記']} ~ {data[-1]['時間戳記']}")
    print(f"   溫度範圍: {min(temps):.1f}°C ~ {max(temps):.1f}°C")
    print(f"   濕度範圍: {min(humis):.1f}% ~ {max(humis):.1f}%")
    print(f"   電燈開啟次數: {lights_on} / {len(data)} ({lights_on/len(data)*100:.1f}%)")
    print()
    
    # 儲存檔案
    print("💾 儲存檔案...")
    save_to_csv(data, 'sensor_data.csv')
    
    if HAS_OPENPYXL:
        save_to_excel(data, 'sensor_data.xlsx')
    
    print()
    print("=" * 60)
    print("✅ 完成！")
    print()
    print("📝 使用方式:")
    print("   1. Flask 版本會自動讀取 sensor_data.csv")
    print("   2. 重新啟動 Flask 應用程式即可看到數據")
    print("   3. 或直接重新整理網頁")
    print()
    print("💡 提示:")
    print("   - 修改腳本中的 count 參數可生成更多數據")
    print("   - 兩個檔案的內容相同，僅格式不同")
    print("=" * 60)

if __name__ == "__main__":
    main()

